# Copyright 2025
# Licensed under the Apache License, Version 2.0

from __future__ import annotations

import io
import os
import zipfile
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ReportLab for enterprise-grade PDF output
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

from .utils import ensure_dir, safe_filename


ProgressCB = Callable[[int, str], None]

OP_CONCAT = "concat_tables"
OP_JOIN = "join_tables"
OP_SHEET_MERGE = "merge_sheets"
OP_SHEET_SPLIT = "split_sheets"
OP_EXCEL_TO_CSV = "excel_to_csv"
OP_CSV_TO_EXCEL = "csv_to_excel"
OP_EXCEL_TO_PDF = "excel_to_pdf"
OP_PDF_TO_EXCEL = "pdf_to_excel"

OP_REGISTRY = {
    OP_CONCAT: "📈 数据表拼接（按列结构一致合并）",
    OP_JOIN: "🔗 多表关联（按 Key 字段 join）",
    OP_SHEET_MERGE: "📑 多文件 Sheet 合并（可选保留样式）",
    OP_SHEET_SPLIT: "✂️ 单文件 Sheet 拆分（ZIP）",
    OP_EXCEL_TO_CSV: "🔄 Excel → CSV（按 Sheet 输出 ZIP）",
    OP_CSV_TO_EXCEL: "📂 CSV → Excel（批量 ZIP）",
    OP_EXCEL_TO_PDF: "📄 Excel → PDF（按 Sheet 输出 ZIP，企业版）",
    OP_PDF_TO_EXCEL: "🔍 PDF → Excel（抽表格，批量 ZIP）",
}


def run_operation(
    operation: str,
    params: Dict[str, Any],
    input_paths: List[Path],
    output_dir: Path,
    progress_cb: Optional[ProgressCB] = None,
) -> Optional[Path]:
    ensure_dir(output_dir)

    def tick(p: int, msg: str) -> None:
        if progress_cb:
            progress_cb(p, msg)

    if operation == OP_CONCAT:
        tick(3, "Reading Excel files...")
        return op_concat_tables(input_paths, output_dir, params, tick)

    if operation == OP_JOIN:
        tick(3, "Reading Excel files...")
        return op_join_tables(input_paths, output_dir, params, tick)

    if operation == OP_SHEET_MERGE:
        tick(3, "Merging sheets...")
        return op_merge_sheets(input_paths, output_dir, params, tick)

    if operation == OP_SHEET_SPLIT:
        tick(3, "Splitting sheets...")
        return op_split_sheets(input_paths, output_dir, params, tick)

    if operation == OP_EXCEL_TO_CSV:
        tick(3, "Converting Excel to CSV...")
        return op_excel_to_csv(input_paths, output_dir, params, tick)

    if operation == OP_CSV_TO_EXCEL:
        tick(3, "Converting CSV to Excel...")
        return op_csv_to_excel(input_paths, output_dir, params, tick)

    if operation == OP_EXCEL_TO_PDF:
        tick(3, "Converting Excel to PDF...")
        return op_excel_to_pdf(input_paths, output_dir, params, tick)

    if operation == OP_PDF_TO_EXCEL:
        tick(3, "Extracting tables from PDF...")
        return op_pdf_to_excel(input_paths, output_dir, params, tick)

    raise ValueError(f"Unknown operation: {operation}")


# ---------------------------
# Operations
# ---------------------------

def _read_all_sheets(xlsx_path: Path) -> List[pd.DataFrame]:
    xl = pd.ExcelFile(xlsx_path)
    dfs = []
    for s in xl.sheet_names:
        dfs.append(pd.read_excel(xlsx_path, sheet_name=s))
    return dfs


def op_concat_tables(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    """
    企业版策略：
    - 读取所有 sheet，按“列名序列”分组
    - 每组输出一个 sheet（避免原版那种“只用第一个结构强行比对”导致丢数据）
    """
    out_name = safe_filename(params.get("output_name", "concat_result.xlsx"))
    out_path = out_dir / out_name

    groups: Dict[str, List[pd.DataFrame]] = {}
    total = max(1, len(input_paths))

    for i, p in enumerate(input_paths, start=1):
        tick(int(5 + 35 * i / total), f"Reading: {p.name}")
        xl = pd.ExcelFile(p)
        for s in xl.sheet_names:
            df = pd.read_excel(p, sheet_name=s)
            sig = "||".join([str(c) for c in df.columns])
            groups.setdefault(sig, []).append(df)

    if not groups:
        # 空输出
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        ws.append(["No tables found"])
        wb.save(out_path)
        tick(90, "No tables found. Output generated.")
        return out_path

    tick(45, "Concatenating groups...")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for idx, (sig, dfs) in enumerate(groups.items(), start=1):
            merged = pd.concat(dfs, ignore_index=True)
            sheet_name = f"Group_{idx}"
            merged.to_excel(writer, index=False, sheet_name=sheet_name)

    tick(95, "Writing Excel...")
    return out_path


def op_join_tables(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    out_name = safe_filename(params.get("output_name", "join_result.xlsx"))
    out_path = out_dir / out_name

    key = params.get("key")
    how = params.get("how", "left")
    if not key:
        raise ValueError("Missing join key")

    # 把所有 sheet 都当作独立表加入 join 候选
    tables: List[pd.DataFrame] = []
    for p in input_paths:
        xl = pd.ExcelFile(p)
        for s in xl.sheet_names:
            df = pd.read_excel(p, sheet_name=s)
            tables.append(df)

    if not tables:
        raise ValueError("No valid tables to join")

    # 过滤不包含 key 的表
    valid = [df for df in tables if key in df.columns]
    if len(valid) < 1:
        raise ValueError(f"No tables contain key column: {key}")

    tick(40, "Joining tables...")
    result = valid[0]
    for i in range(1, len(valid)):
        tick(40 + int(40 * i / max(1, len(valid) - 1)), f"Joining {i+1}/{len(valid)}")
        result = pd.merge(
            result,
            valid[i],
            on=key,
            how=how,
            suffixes=("", f"_t{i+1}"),
        )

    tick(90, "Writing Excel...")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="Joined")

    return out_path


def op_merge_sheets(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    out_name = safe_filename(params.get("output_name", "merged_sheets.xlsx"))
    out_path = out_dir / out_name
    preserve_styles = bool(params.get("preserve_styles", False))

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    total_files = max(1, len(input_paths))
    sheet_counter = 0

    for fi, p in enumerate(input_paths, start=1):
        tick(int(5 + 70 * fi / total_files), f"Merging: {p.name}")

        if preserve_styles:
            wb_in = load_workbook(p, data_only=False)
            for ws_in in wb_in.worksheets:
                title = ws_in.title
                new_title = _unique_sheet_name(wb_out, title)
                ws_out = wb_out.create_sheet(new_title)
                _copy_sheet_with_styles(ws_in, ws_out)
                sheet_counter += 1
        else:
            xl = pd.ExcelFile(p)
            for s in xl.sheet_names:
                df = pd.read_excel(p, sheet_name=s)
                new_title = _unique_sheet_name(wb_out, s)
                ws = wb_out.create_sheet(new_title)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                sheet_counter += 1

    tick(90, f"Saving workbook ({sheet_counter} sheets)...")
    wb_out.save(out_path)
    return out_path


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    base = base[:31]
    name = base
    k = 1
    existing = {ws.title for ws in wb.worksheets}
    while name in existing:
        suffix = f"_{k}"
        name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
        k += 1
    return name


def _copy_sheet_with_styles(ws_in, ws_out) -> None:
    # 复制列宽/行高
    for col, dim in ws_in.column_dimensions.items():
        ws_out.column_dimensions[col].width = dim.width
    for row, dim in ws_in.row_dimensions.items():
        ws_out.row_dimensions[row].height = dim.height

    # 复制单元格
    for row in ws_in.iter_rows():
        for cell in row:
            new_cell = ws_out.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = cell._style
            new_cell.number_format = cell.number_format
            new_cell.alignment = cell.alignment
            new_cell.font = cell.font
            new_cell.border = cell.border
            new_cell.fill = cell.fill
            new_cell.protection = cell.protection

    # 合并单元格
    for merged in ws_in.merged_cells.ranges:
        ws_out.merge_cells(str(merged))


def op_split_sheets(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    if len(input_paths) != 1:
        raise ValueError("Sheet splitting requires exactly 1 Excel file")

    xlsx = input_paths[0]
    base = xlsx.stem
    out_zip = out_dir / safe_filename(params.get("output_name", f"{base}_split.zip"))

    xl = pd.ExcelFile(xlsx)
    names = xl.sheet_names
    if not names:
        raise ValueError("No sheets found")

    tick(20, "Preparing ZIP...")
    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for i, s in enumerate(names, start=1):
            tick(20 + int(70 * i / len(names)), f"Exporting sheet: {s}")
            df = pd.read_excel(xlsx, sheet_name=s)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=s[:31])
            buf.seek(0)
            z.writestr(f"{safe_filename(base)}_{safe_filename(s)}.xlsx", buf.getvalue())

    tick(95, "ZIP ready")
    return out_zip


def op_excel_to_csv(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    out_zip = out_dir / safe_filename(params.get("output_name", "excel_to_csv.zip"))
    total = max(1, len(input_paths))

    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for fi, p in enumerate(input_paths, start=1):
            tick(int(5 + 80 * fi / total), f"Processing: {p.name}")
            xl = pd.ExcelFile(p)
            for s in xl.sheet_names:
                df = pd.read_excel(p, sheet_name=s)
                csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
                z.writestr(f"{p.stem}_{safe_filename(s)}.csv", csv_bytes)

    tick(95, "ZIP ready")
    return out_zip


def op_csv_to_excel(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    out_zip = out_dir / safe_filename(params.get("output_name", "csv_to_excel.zip"))
    total = max(1, len(input_paths))

    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for i, p in enumerate(input_paths, start=1):
            tick(int(5 + 85 * i / total), f"Reading CSV: {p.name}")
            df = _read_csv_with_fallback(p)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
            buf.seek(0)
            z.writestr(f"{p.stem}.xlsx", buf.getvalue())

    tick(95, "ZIP ready")
    return out_zip


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    # 本地企业常见：utf-8 / utf-8-sig / gbk
    for enc in (None, "utf-8-sig", "gbk", "latin-1"):
        try:
            if enc is None:
                return pd.read_csv(path)
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, encoding="utf-8", encoding_errors="ignore")


def op_excel_to_pdf(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    """
    企业版 PDF：
    - 用 ReportLab 输出表格
    - 用内置 UnicodeCIDFont 支持中文（STSong-Light）
    - 每个 sheet 一个 pdf，最终 ZIP
    """
    out_zip = out_dir / safe_filename(params.get("output_name", "excel_to_pdf.zip"))
    page_mode = params.get("page_mode", "landscape")  # landscape / portrait
    max_rows = int(params.get("max_rows", 200))       # 防止超大表格 PDF 爆炸

    # register Chinese font (CID)
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        base_font = "STSong-Light"
    except Exception:
        base_font = "Helvetica"

    total = max(1, len(input_paths))
    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for fi, p in enumerate(input_paths, start=1):
            tick(int(5 + 80 * fi / total), f"Rendering PDFs: {p.name}")
            xl = pd.ExcelFile(p)
            for s in xl.sheet_names:
                df = pd.read_excel(p, sheet_name=s)
                if len(df) > max_rows:
                    df = df.head(max_rows)

                pdf_buf = io.BytesIO()
                _df_to_pdf(
                    df=df,
                    title=f"{p.stem} / {s}",
                    out_stream=pdf_buf,
                    font_name=base_font,
                    landscape_mode=(page_mode == "landscape"),
                )
                pdf_buf.seek(0)
                z.writestr(f"{p.stem}_{safe_filename(s)}.pdf", pdf_buf.getvalue())

    tick(95, "ZIP ready")
    return out_zip


def _df_to_pdf(df: pd.DataFrame, title: str, out_stream: io.BytesIO, font_name: str, landscape_mode: bool) -> None:
    styles = getSampleStyleSheet()
    page_size = landscape(A4) if landscape_mode else A4

    doc = SimpleDocTemplate(out_stream, pagesize=page_size, leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    story = []

    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 8))

    # Build table data (header + rows)
    header = [str(c) for c in df.columns]
    body = df.astype(str).values.tolist()
    data = [header] + body

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ]
        )
    )

    story.append(tbl)
    doc.build(story)


def op_pdf_to_excel(input_paths: List[Path], out_dir: Path, params: Dict[str, Any], tick: ProgressCB) -> Path:
    out_zip = out_dir / safe_filename(params.get("output_name", "pdf_to_excel.zip"))
    total = max(1, len(input_paths))

    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for i, p in enumerate(input_paths, start=1):
            tick(int(5 + 85 * i / total), f"Extracting: {p.name}")
            all_tables: List[pd.DataFrame] = []

            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables() or []
                    for t in tables:
                        if not t or len(t) < 2:
                            continue
                        df = pd.DataFrame(t[1:], columns=t[0])
                        all_tables.append(df)

            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                if all_tables:
                    for k, df in enumerate(all_tables, start=1):
                        sheet = f"Table_{k}"
                        writer_sheet = sheet[:31]
                        df.to_excel(writer, index=False, sheet_name=writer_sheet)
                else:
                    pd.DataFrame([{"Message": "No tables found"}]).to_excel(writer, index=False, sheet_name="No_Tables")

            buf.seek(0)
            z.writestr(f"{p.stem}_tables.xlsx", buf.getvalue())

    tick(95, "ZIP ready")
    return out_zip
